from openpyxl import Workbook
import get_document

def export_excel(corp_name, document, report_nm):
    if type(document) == " ":
        print()
        print(report_nm + "의 내용은 현재 삭제되었습니다")
        print()
        return 0
    wb = Workbook()
    xlsx = wb.active
    title_name = corp_name + "-" + report_nm.replace("[", "").replace("]", '-')
    xlsx.title = title_name
    table = document[0].find_all('table')
    title = []
    row_index = 1
    column_index = 1
    name = document[0].find('title').get_text().split()
    xlsx.cell(row=row_index, column=column_index).value = name[1]
    row_index = row_index + 2
    for tb in table:
        for tr in tb.find_all('tr'):
            if len(tr.find_all('th')) != 0:
                column_index = column_index + 1
            for p in tr.find_all('p'):
                num = p.get_text().replace(",", "")
                num = num.replace("(", "")
                num = num.replace(")", "")
                if num.isdigit():
                    num = int(num)
                xlsx.cell(row=row_index, column=column_index).value = num
                column_index = column_index + 1
            row_index = row_index + 1
            column_index = 1

        row_index = row_index + 1
    wb.save(filename=title_name + '.xlsx')

# export_excel('삼천당제약',get_document.get_document('20180618000112','[기재정정]사업보고서 (2017.12)'),'[기재정정]사업보고서 (2017.12)')